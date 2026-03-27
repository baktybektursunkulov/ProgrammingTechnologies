import json
from pathlib import Path

# Базовый путь к sample_data
BASE = Path(__file__).parent / "sample_data"

# ==================== Задача 1 ====================
def task1_json_comprehensions_json():
    """JSON → Comprehensions → JSON"""
    with open(BASE / "students.json", encoding="utf-8") as f:
        students = json.load(f)

    young_students = [s["name"] for s in students if s["age"] < 22]
    astana_grades = {s["name"]: s["grade"] for s in students if s["city"] == "Astana"}

    result = {"young_students": young_students, "astana_grades": astana_grades}
    with open(BASE / "filtered_students.json", "w", encoding="utf-8") as f:
        json.dump(result, f, indent=2, ensure_ascii=False)

    print("Задача 1: filtered_students.json создан")
    return result

# ==================== Задача 2 ====================
class JsonLinesIterator:
    """Итератор для построчного чтения JSON Lines (.jsonl)."""

    def __init__(self, file_path):
        self.file_path = Path(file_path)
        self._file = None

    def __iter__(self):
        self._file = open(self.file_path, encoding="utf-8")
        return self

    def __next__(self):
        if self._file is None:
            raise StopIteration
        line = self._file.readline()
        if not line:
            self._file.close()
            self._file = None
            raise StopIteration
        return json.loads(line.strip())

    def __enter__(self):
        return self

    def __exit__(self, *args):
        if self._file:
            self._file.close()


def task2_jsonl_iterator():
    """Итератор по JSONL файлу"""
    print("\nЗадача 2: JsonLinesIterator")
    for record in JsonLinesIterator(BASE / "logs.jsonl"):
        print(f"  {record}")


# ==================== Задача 3 ====================
def task3_xlsx_to_json():
    """XLSX → Comprehensions → JSON"""
    from openpyxl import load_workbook

    wb = load_workbook(BASE / "sales.xlsx")
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    rows = list(ws.iter_rows(min_row=2, values_only=True))

    sales = [dict(zip(headers, row)) for row in rows]

    large_orders = [s for s in sales if s["quantity"] >= 15]
    products = {s["product"] for s in sales}
    quantity_by_product = {
        p: sum(s["quantity"] for s in sales if s["product"] == p)
        for p in products
    }

    result = {"large_orders": large_orders, "quantity_by_product": quantity_by_product}
    with open(BASE / "sales_summary.json", "w", encoding="utf-8") as f:
        json.dump(result, f, indent=2, ensure_ascii=False)

    print("\nЗадача 3: sales_summary.json создан")
    return result


# ==================== Задача 4 ====================
def task4_products_to_xlsx():
    """JSON + Set comprehension + XLSX"""
    from openpyxl import Workbook

    with open(BASE / "products.json", encoding="utf-8") as f:
        products = json.load(f)

    categories = {p["category"] for p in products}
    expensive = [p for p in products if p["price"] > 1000]

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Categories"
    for i, cat in enumerate(sorted(categories), 1):
        ws1.cell(row=i, column=1, value=cat)

    ws2 = wb.create_sheet("Expensive Products")
    ws2.append(["name", "category", "price"])
    for p in expensive:
        ws2.append([p["name"], p["category"], p["price"]])

    wb.save(BASE / "products_by_category.xlsx")
    print("\nЗадача 4: products_by_category.xlsx создан")


# ==================== Задача 5 ====================
def read_json_chunks(file_path, chunk_size=5):
    """Генератор: возвращает по chunk_size элементов из JSON-массива."""
    with open(file_path, encoding="utf-8") as f:
        data = json.load(f)
    for i in range(0, len(data), chunk_size):
        yield data[i : i + chunk_size]


def task5_generator_json():
    """Генератор + JSON"""
    print("\nЗадача 5: read_json_chunks")
    for chunk in read_json_chunks(BASE / "students.json", chunk_size=2):
        names = [s["name"] for s in chunk]
        print(f"  Чанк: {names}")
        name_grade = {s["name"]: s["grade"] for s in chunk}
        print(f"  Dict: {name_grade}")


# ==================== Задача 6 ====================
def students_to_xlsx(students, output_path):
    """Конвертация списка студентов в XLSX."""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    headers = ["№", "name", "age", "grade", "city"]
    ws.append(headers)
    for i, s in enumerate(students, 1):
        row = [i] + [s.get(h) for h in headers[1:]]
        ws.append(row)
    wb.save(output_path)


def xlsx_to_students(input_path):
    """Чтение XLSX и конвертация в список словарей."""
    from openpyxl import load_workbook

    wb = load_workbook(input_path)
    ws = wb.active
    headers = [cell.value for cell in ws[1] if cell.value != "№"]
    students = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        students.append(dict(zip(headers, row[1:])))
    return students


def task6_json_xlsx_roundtrip():
    """JSON ↔ XLSX с итераторами"""
    with open(BASE / "students.json", encoding="utf-8") as f:
        original = json.load(f)

    xlsx_path = BASE / "students_export.xlsx"
    students_to_xlsx(original, xlsx_path)
    restored = xlsx_to_students(xlsx_path)

    print("\nЗадача 6: JSON → XLSX → JSON")
    print(f"  Исходных записей: {len(original)}")
    print(f"  Восстановлено: {len(restored)}")
    print(f"  Совпадают: {original == restored}")


# ==================== Задача 7 ====================
def task7_nested_comprehensions():
    """Извлечение тегов из вложенных структур"""
    with open(BASE / "products.json", encoding="utf-8") as f:
        products = json.load(f)

    tags = {tag for p in products for tag in p.get("tags", [])}
    result = sorted(tags)

    with open(BASE / "extracted_tags.json", "w", encoding="utf-8") as f:
        json.dump(result, f, indent=2, ensure_ascii=False)

    print("\nЗадача 7: extracted_tags.json создан")
    return result


# ==================== Main ====================
if __name__ == "__main__":
    # task1_json_comprehensions_json()
    # task2_jsonl_iterator()
    # task3_xlsx_to_json()
    task4_products_to_xlsx()
    # task5_generator_json()
    # task6_json_xlsx_roundtrip()
    # task7_nested_comprehensions()
    print("\nВсе задачи выполнены.")
