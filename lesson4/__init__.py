# Lesson 5: Iterators, Comprehensions, JSON, XLSX
import json
#1
def task1():

    with open( "sample_data/students.json", encoding="utf-8") as f:
        students = json.load(f)
    top_students = [x['name'] for x in students if x['grade']>80]
    print(top_students)
    almaty_ages = {x['name']: x['age'] for x in students if x['city']=='Almaty'}
    print(almaty_ages)
    result  = {}
    result['top_students'] = top_students
    result['almaty_ages'] = almaty_ages
    with open("sample_data/filtered_students1.json", "w", encoding="utf-8") as f:
        json.dump(result, f, indent=2, ensure_ascii=False)



def task3():
    from openpyxl import load_workbook

    wb = load_workbook("sample_data/sales.xlsx")
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    sales = [dict(zip(headers, row)) for row in rows]

    moreThanTenProducts = [s['product'] for s in sales if s['quantity']>10]

    product = {s['product']: s['price']*s['quantity'] for s in sales}
    result = {}
    result['moreThanTenProducts'] = moreThanTenProducts
    result['product'] = product

    with open("sample_data/sales_summary.json", "w" , encoding="utf-8") as f:
        json.dump(result, f, indent=2, ensure_ascii=False)

if __name__ == "__main__":
    # task1()
    task3()